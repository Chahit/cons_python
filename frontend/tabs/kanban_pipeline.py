import streamlit as st
import pandas as pd
import numpy as np
import sys, os
import io
from datetime import date, timedelta
import calendar as _cal

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from styles import page_header, skeleton_loader

# ── Period preset shortcuts ────────────────────────────────────────────────
# Used as quick-select chips in the calendar date picker.
_PERIOD_PRESETS = [
    {"label": "15 Days",    "days": 15},
    {"label": "1 Month",    "days": 30},
    {"label": "2 Months",   "days": 60},
    {"label": "3 Months",   "days": 90},
    {"label": "6 Months",   "days": 180},
    {"label": "1 Year",     "days": 365},
]
# Legacy config kept for fallback / export labeling
_PERIOD_CONFIG = {
    "Monthly":    {"multiplier": 1/3,  "days": 30,  "label": "Monthly"},
    "Quarterly":  {"multiplier": 1.0,  "days": 90,  "label": "Quarterly (90d)"},
    "Yearly":     {"multiplier": 4.0,  "days": 365, "label": "Annual"},
}

# ── Kanban swimlane configuration ──────────────────────────────────────────
# Industry-standard partner tiers (Salesforce/Gartner alignment)
# Each lane now includes a playbook: the recommended sales action for that tier.
LANES = [
    {
        "key": "champion",
        "label": "🏆 Champion",
        "segments": {"Champion"},
        "color": "#22c55e",
        "playbook": "Schedule QBR · Offer loyalty pricing · Co-marketing opportunities",
        "tier": "Gold",
    },
    {
        "key": "emerging",
        "label": "🚀 Emerging",
        "segments": {"Emerging"},
        "color": "#06b6d4",
        "playbook": "Onboard new categories · Assign growth rep · Bundle incentives",
        "tier": "Silver",
    },
    {
        "key": "healthy",
        "label": "✅ Healthy",
        "segments": {"Healthy"},
        "color": "#3b82f6",
        "playbook": "Maintain cadence · Upsell adjacent categories · Regular check-ins",
        "tier": "Bronze",
    },
    {
        "key": "at_risk",
        "label": "⚠️ At Risk",
        "segments": {"At Risk"},
        "color": "#f59e0b",
        "playbook": "Retention call within 7 days · Diagnose root cause · Win-back offer",
        "tier": "Watch",
    },
    {
        "key": "critical",
        "label": "🔴 Critical",
        "segments": {"Critical"},
        "color": "#ef4444",
        "playbook": "Immediate escalation · Executive outreach · Last-resort pricing",
        "tier": "Alert",
    },
]

def _fmt_inr(val):
    try:
        v = float(val)
    except Exception:
        return "₹0"
    if v >= 1_00_00_000: return f"₹{v / 1_00_00_000:.1f}Cr"
    if v >= 1_00_000:    return f"₹{v / 1_00_000:.1f}L"
    if v >= 1_000:       return f"₹{v / 1_000:.0f}K"
    return f"₹{v:.0f}"

def _days_ago(recency_days):
    """Format recency_days as human-readable last-order string."""
    try:
        d = int(float(recency_days))
        if d == 0:  return "Today"
        if d == 1:  return "1d ago"
        if d < 31:  return f"{d}d ago"
        if d < 365: return f"{d // 30}mo ago"
        return f"{d // 365}y ago"
    except Exception:
        return "—"

def _primary_risk_signal(row):
    """Return (signal_text, color) for the single most urgent risk indicator."""
    try:
        recency   = float(row.get("recency_days") or 0)
        drop      = float(row.get("revenue_drop_pct") or 0)
        recent    = float(row.get("recent_90_revenue") or 0)
        prev      = float(row.get("prev_90_revenue") or 0)
        if recent == 0 and prev > 0:
            return "🔴 Revenue stopped", "#ef4444"
        if recency > 120:
            return f"⏰ {int(recency)}d no purchase", "#ef4444"
        if drop > 50:
            return f"📉 Revenue ↓{drop:.0f}%", "#ef4444"
        if recency > 60:
            return f"⏳ {int(recency)}d quiet", "#f59e0b"
        if drop > 25:
            return f"📉 ↓{drop:.0f}% revenue", "#f59e0b"
    except Exception:
        pass
    return None, None

# ── Cache the heavy data slice so repeated renders are fast ────────────────
@st.cache_data(ttl=300, show_spinner=False)
def _build_kanban_df(_pf_df, _df_hash: str = ""):
    """Extract and pre-process only the columns needed by the Kanban board.
    _df_hash is a scalar derived from the index so @st.cache_data keys correctly.
    """
    needed = [
        "company_name", "state", "health_segment", "health_status",
        "health_score", "growth_rate_90d",
        "churn_probability", "credit_risk_band",
        "recent_90_revenue", "prev_90_revenue", "revenue_drop_pct", "revenue_at_risk",
        "cluster_label", "cluster_type", "recency_days",
    ]
    df = _pf_df.reset_index()
    if "company_name" not in df.columns and "index" in df.columns:
        df = df.rename(columns={"index": "company_name"})
    cols_present = [c for c in needed if c in df.columns]
    df = df[cols_present].copy()

    # Fill defaults for optional columns
    for col, default in [
        ("churn_probability", 0.0),
        ("credit_risk_band", "N/A"),
        ("health_segment", "Healthy"),
        ("health_score", 0.5),
        ("growth_rate_90d", 0.0),
        ("cluster_label", "—"),
        ("cluster_type", "—"),
        ("recency_days", None),
        ("prev_90_revenue", 0.0),
        ("revenue_drop_pct", None),
        ("revenue_at_risk", None),
    ]:
        if col not in df.columns:
            df[col] = default


    df["churn_probability"] = pd.to_numeric(df["churn_probability"], errors="coerce").fillna(0.0)
    df["recent_90_revenue"] = pd.to_numeric(df.get("recent_90_revenue", 0), errors="coerce").fillna(0.0)

    # Compute revenue_at_risk if not provided: churn probability × 90d revenue
    # Fix: if a partner has never ordered (recent_90_revenue=0) but churn is high,
    # use a minimum floor of Rs 5K/month × 3 (90d) so they don't show as 0 risk.
    _MIN_FLOOR_90D = 15_000  # Rs 15K as absolute floor for churn >=0.5 partners
    if "revenue_at_risk" not in df.columns or df["revenue_at_risk"].isna().all():
        _base = df["churn_probability"] * df["recent_90_revenue"]
        _floor_mask = (df["recent_90_revenue"] == 0) & (df["churn_probability"] >= 0.5)
        _base = _base.where(~_floor_mask, df["churn_probability"] * _MIN_FLOOR_90D)
        df["revenue_at_risk"] = _base
    else:
        df["revenue_at_risk"] = pd.to_numeric(df["revenue_at_risk"], errors="coerce")
        _base_fallback = df["churn_probability"] * df["recent_90_revenue"]
        _floor_mask = (df["recent_90_revenue"] == 0) & (df["churn_probability"] >= 0.5)
        _base_fallback = _base_fallback.where(~_floor_mask, df["churn_probability"] * _MIN_FLOOR_90D)
        df["revenue_at_risk"] = df["revenue_at_risk"].fillna(_base_fallback)

    return df


@st.cache_data(ttl=300, show_spinner=False)
def _build_category_performance(_group_spend_df):
    """
    Build a category→partner performance summary.
    Returns a DataFrame: company_name, category (group_name), total_spend
    """
    if _group_spend_df is None or _group_spend_df.empty:
        return pd.DataFrame(columns=["company_name", "category", "total_spend"])
    df = _group_spend_df.copy()
    if "group_name" in df.columns:
        df = df.rename(columns={"group_name": "category"})
    needed = [c for c in ["company_name", "category", "total_spend"] if c in df.columns]
    return df[needed].copy()


# ── Period-calibrated thresholds ──────────────────────────────────────────
# For each time period: (churn_max, drop_max_champion, drop_max_healthy,
#                        score_min, rev_tier_pct, rev_tier_churn_cap,
#                        rev_tier_drop_cap, abs_rev_floor_scaled)
# abs_rev_floor_scaled: absolute annual equiv threshold; partners above this
# get a softer degrowth cap because they define the core business
_PERIOD_THRESHOLDS = {
    # Monthly (1-month window — noisy; be lenient, especially for high-rev)
    "Monthly": dict(
        score_champ=0.68, drop_champ=12.0, churn_champ=0.40,
        score_healthy=0.48, drop_healthy=18.0,
        rev_tier_pct=0.88, rev_tier_churn=0.55, rev_tier_drop=30.0, rev_tier_score=0.38,
        abs_floor_yr=1_00_00_000,   # ₹1Cr/yr = ₹8L/mo — top accounts
        abs_churn_cap=0.58, abs_drop_cap=35.0,
        dg_thresh=18.0,
    ),
    # Quarterly (90-day actual — model ground truth)
    "Quarterly": dict(
        score_champ=0.72, drop_champ=10.0, churn_champ=0.35,
        score_healthy=0.50, drop_healthy=20.0,
        rev_tier_pct=0.90, rev_tier_churn=0.50, rev_tier_drop=25.0, rev_tier_score=0.40,
        abs_floor_yr=3_00_00_000,   # ₹3Cr/yr = ₹75L/quarter
        abs_churn_cap=0.55, abs_drop_cap=30.0,
        dg_thresh=20.0,
    ),
    # Yearly (4× projection — forward-looking; broader tolerances)
    "Yearly": dict(
        score_champ=0.68, drop_champ=15.0, churn_champ=0.40,
        score_healthy=0.46, drop_healthy=25.0,
        rev_tier_pct=0.85, rev_tier_churn=0.55, rev_tier_drop=32.0, rev_tier_score=0.38,
        abs_floor_yr=5_00_00_000,   # ₹5Cr/yr — key accounts
        abs_churn_cap=0.58, abs_drop_cap=35.0,
        dg_thresh=25.0,
    ),
}


def _recompute_segments_for_period(
    df: pd.DataFrame, rev_mult: float, period: str = "Quarterly"
) -> pd.DataFrame:
    """
    Re-classify health_segment based on period-scaled revenue with
    precision-calibrated thresholds per time window.

    Champion paths (in priority order):
      1. Lifetime Revenue Shield  — top-10% LTV still buying, churn < 0.50
      2. Score path               — high composite score + low churn
      3. Revenue-tier path        — top % by scaled rev, not truly churning
      4. Absolute revenue floor   — ₹-floor partners get softer caps
    """
    if df.empty or "recent_90_revenue" not in df.columns:
        return df

    th = _PERIOD_THRESHOLDS.get(period, _PERIOD_THRESHOLDS["Quarterly"])
    df = df.copy()

    scaled_rev  = df["recent_90_revenue"] * rev_mult
    churn_p     = pd.to_numeric(df.get("churn_probability",  pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    drop_pct    = pd.to_numeric(df.get("revenue_drop_pct",  pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    prev_rev    = pd.to_numeric(df.get("prev_90_revenue",   pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0) * rev_mult
    score       = pd.to_numeric(df.get("health_score",      pd.Series(0.5, index=df.index)), errors="coerce").fillna(0.5)
    growth      = pd.to_numeric(df.get("growth_rate_90d",   pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    recency     = pd.to_numeric(df.get("recency_days",      pd.Series(0,   index=df.index)), errors="coerce").fillna(0)
    lifetime    = pd.to_numeric(df.get("lifetime_revenue",  pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    recent_txns = pd.to_numeric(df.get("recent_txns",       pd.Series(0,   index=df.index)), errors="coerce").fillna(0)

    rev_tier_pct     = th["rev_tier_pct"]
    rev_tier_q       = float(scaled_rev.quantile(rev_tier_pct)) if len(scaled_rev) > 0 else 0.0
    ltv_p90          = float(lifetime.quantile(0.90)) if len(lifetime) > 0 else 0.0
    abs_floor_scaled = th["abs_floor_yr"] * rev_mult / 4.0
    recency_ok       = recency <= 120

    segments        = []
    gone_quiet_list = []

    for i in df.index:
        sr   = float(scaled_rev.get(i, 0)   or 0)
        pr   = float(prev_rev.get(i, 0)     or 0)
        cp   = float(churn_p.get(i, 0)      or 0)
        dp   = float(drop_pct.get(i, 0)     or 0)
        hs   = float(score.get(i, 0.5)      or 0.5)
        gr   = float(growth.get(i, 0)       or 0)
        rok  = bool(recency_ok.get(i, True))
        ltv  = float(lifetime.get(i, 0)     or 0)
        txns = float(recent_txns.get(i, 0)  or 0)

        is_top10_ltv = ltv >= ltv_p90 and ltv_p90 > 0
        is_abs_floor = sr  >= abs_floor_scaled and abs_floor_scaled > 0

        # Industry Silver tier: growth velocity gate.
        # Requires positive QoQ (>=5%) AND >= 2 purchases to prevent
        # single-order false positives. Stable + highly engaged also qualifies.
        growth_velocity_ok = (
            sr > 0 and cp < 0.55 and hs < th["score_champ"] and rok
            and (
                (gr >= 0.05 and txns >= 2)
                or (gr >= -0.05 and txns >= 3 and cp < 0.30)
            )
        )

        # Gone Quiet: had prior revenue but zero in current window
        # (boundary edge case — not confirmed churn yet, shown as UI badge)
        gone_quiet = sr <= 0 and pr > 0
        gone_quiet_list.append(gone_quiet)

        # ❛ Revenue stopped after prior activity -> confirmed churn
        if sr <= 0 and pr > 0:
            segments.append("Critical")
        # ❜ Lifetime Revenue Shield
        elif is_top10_ltv and sr > 0 and cp < 0.50 and dp < 30.0:
            segments.append("Champion")
        # ❝ Score path
        elif hs >= th["score_champ"] and dp < th["drop_champ"] and cp < th["churn_champ"] and rok:
            segments.append("Champion")
        # ❞ Revenue-tier Champion
        elif sr >= rev_tier_q and sr > 0 and cp < th["rev_tier_churn"] and dp < th["rev_tier_drop"] and hs >= th["rev_tier_score"] and rok:
            segments.append("Champion")
        # ❟ Absolute revenue floor
        elif is_abs_floor and sr > 0 and cp < th["abs_churn_cap"] and dp < th["abs_drop_cap"] and rok:
            segments.append("Champion")
        # ❠ Emerging: growth velocity gate (industry Silver tier)
        elif growth_velocity_ok:
            segments.append("Emerging")
        # ❡ Healthy
        elif hs >= th["score_healthy"] and dp < th["drop_healthy"] and rok:
            segments.append("Healthy")
        # ❢ At Risk
        elif hs >= 0.30:
            segments.append("At Risk")
        # ❣ Critical
        else:
            segments.append("Critical")

    df["health_segment"] = segments
    df["_gone_quiet"]    = gone_quiet_list
    return df


@st.cache_data(ttl=600, show_spinner=False)
def _fetch_custom_period_revenue(_engine, start_dt: date, end_dt: date) -> pd.DataFrame:
    """
    Query the DB for actual partner revenue within [start_dt, end_dt].
    Returns a DataFrame with columns:
        company_name, period_revenue, period_txns, last_purchase_date
    Falls back to an empty DataFrame if the DB is unreachable or
    transactions_dsr is not available.
    """
    if _engine is None:
        return pd.DataFrame()
    query = """
    SELECT
        mp.company_name,
        COALESCE(SUM(tp.net_amt), 0)          AS period_revenue,
        COUNT(DISTINCT t.id)                  AS period_txns,
        MAX(t.date)::date                     AS last_purchase_date
    FROM transactions_dsr t
    JOIN transactions_dsr_products tp ON t.id = tp.dsr_id
    JOIN master_party mp ON t.party_id = mp.id
    WHERE LOWER(CAST(t.is_approved AS TEXT)) = 'true'
      AND t.date >= '{start}'::date
      AND t.date <= '{end}'::date
    GROUP BY mp.company_name
    """.format(start=start_dt.isoformat(), end=end_dt.isoformat())
    try:
        df = pd.read_sql(query, _engine)
        df["last_purchase_date"] = pd.to_datetime(df["last_purchase_date"], errors="coerce")
        return df
    except Exception:
        return pd.DataFrame()



# ── Temporal segment migration helpers ────────────────────────────────────

def _derive_change_reason(row) -> str:
    """
    Return a human-readable, deterministic explanation for why a partner
    changed health segment between two periods.
    Priority: revenue stop > recency gap > revenue drop > churn spike > growth > score.
    """
    sr       = float(row.get("period_rev_curr", 0) or 0)
    pr       = float(row.get("period_rev_prev", 0) or 0)
    recency  = float(row.get("recency_days",    9999) or 9999)
    drop     = float(row.get("revenue_drop_pct",   0) or 0)
    churn    = float(row.get("churn_probability",   0) or 0)
    gr       = float(row.get("growth_rate_90d",     0) or 0)
    hs_curr  = float(row.get("health_score",       0.5) or 0.5)
    seg_prev = str(row.get("segment_prev", ""))
    seg_curr = str(row.get("segment_curr", ""))

    _tier = ["Critical", "At Risk", "Healthy", "Emerging", "Champion"]
    try:
        improved = _tier.index(seg_curr) > _tier.index(seg_prev)
    except ValueError:
        improved = False

    if sr <= 0 and pr > 0:
        return "Revenue dropped to Rs 0 — no purchases in current period (churned)"
    if recency > 120:
        return f"No purchase in {int(recency)}d — account went quiet"
    if drop > 40:
        return f"Revenue fell {drop:.0f}% vs prior period"
    if churn > 0.70:
        return f"Churn risk spiked to {churn:.0%} (AI flag — high urgency)"
    if improved and gr >= 0.20:
        return f"Revenue grew {gr*100:.0f}% QoQ — rising star signal"
    if improved and gr >= 0.05:
        return f"Consistent growth ({gr*100:.0f}% QoQ) lifted score"
    if drop > 20:
        return f"Revenue fell {drop:.0f}% — degrowth threshold crossed"
    if churn > 0.45:
        return f"Churn probability rose to {churn:.0%}"
    if improved:
        return f"Health score improved to {hs_curr:.2f} — tier boundary crossed"
    return f"Health score declined to {hs_curr:.2f} — tier boundary crossed"


def _compute_prior_window_segments(
    base_df: pd.DataFrame,
    engine,
    sel_start: date,
    span_days: int,
    rev_mult: float,
) -> pd.DataFrame:
    """
    Fetch revenue for the prior equivalent window (same span, just before sel_start)
    and recompute health segments. Returns a DataFrame with columns:
        company_name, segment_prior, period_rev_prior
    """
    prior_end   = sel_start - timedelta(days=1)
    prior_start = sel_start - timedelta(days=span_days)

    # Use session_state cache to avoid repeat DB hits
    _key = f"rev_{prior_start}_{prior_end}"
    if _key in st.session_state:
        prior_rev_df = st.session_state[_key]
    else:
        prior_rev_df = _fetch_custom_period_revenue(engine, prior_start, prior_end)
        st.session_state[_key] = prior_rev_df

    df_prior = base_df.copy()
    if not prior_rev_df.empty and "company_name" in prior_rev_df.columns:
        df_prior = df_prior.merge(
            prior_rev_df[["company_name", "period_revenue"]],
            on="company_name", how="left",
        )
        df_prior["period_revenue"] = pd.to_numeric(df_prior["period_revenue"], errors="coerce").fillna(0.0)
        df_prior["recent_90_revenue"] = df_prior["period_revenue"]
        df_prior = df_prior.drop(columns=["period_revenue"], errors="ignore")
    else:
        df_prior["recent_90_revenue"] = df_prior["recent_90_revenue"] * (span_days / 90.0)

    df_prior = _recompute_segments_for_period(df_prior, rev_mult, period="Quarterly")
    return df_prior[["company_name", "health_segment", "recent_90_revenue"]].rename(
        columns={"health_segment": "segment_prior", "recent_90_revenue": "period_rev_prior"}
    )


def _render_segment_migration(
    df_curr: pd.DataFrame,
    base_df: pd.DataFrame,
    engine,
    sel_start: date,
    sel_end: date,
    span_days: int,
    rev_mult: float,
):
    """
    Render the Before → After segment migration panel:
    1. Transition matrix (heat-map table)
    2. Expandable movers list with human-readable change reasons
    """
    SEG_ORDER  = ["Champion", "Emerging", "Healthy", "At Risk", "Critical"]
    SEG_COLORS = {
        "Champion": "#22c55e", "Emerging": "#06b6d4", "Healthy": "#3b82f6",
        "At Risk":  "#f59e0b", "Critical": "#ef4444",
    }

    prior_start = sel_start - timedelta(days=span_days)
    prior_end   = sel_start - timedelta(days=1)

    st.markdown(
        f"""
        <div style='background:#0f1117;border:1px solid #1e3a5f;border-radius:12px;
             padding:14px 20px;margin-bottom:12px;'>
          <div style='font-size:13px;font-weight:800;color:#93c5fd;margin-bottom:4px;'>
            ⏳ Segment Migration Analysis
          </div>
          <div style='font-size:12px;color:#64748b;'>
            Prior window: <b style='color:#7eb8f0;'>{prior_start.strftime('%d %b %Y')}
            &rarr; {prior_end.strftime('%d %b %Y')}</b>
            &nbsp;→&nbsp;
            Current: <b style='color:#7eb8f0;'>{sel_start.strftime('%d %b %Y')}
            &rarr; {sel_end.strftime('%d %b %Y')}</b>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.spinner("Computing prior-period segments..."):
        df_prior = _compute_prior_window_segments(base_df, engine, sel_start, span_days, rev_mult)

    # Merge current + prior
    curr_slim = df_curr[["company_name", "health_segment", "recent_90_revenue",
                          "churn_probability", "revenue_drop_pct",
                          "growth_rate_90d", "health_score", "recency_days"]].copy()
    curr_slim = curr_slim.rename(columns={"health_segment": "segment_curr",
                                          "recent_90_revenue": "period_rev_curr"})
    merged = curr_slim.merge(df_prior, on="company_name", how="inner")
    if merged.empty:
        st.info("No common partners between the two windows.")
        return

    changers = merged[merged["segment_curr"] != merged["segment_prev"]].copy() if "segment_prev" in merged.columns else merged[merged["segment_curr"] != merged["segment_prior"]].copy()
    # Normalise column name
    if "segment_prior" in merged.columns and "segment_prev" not in merged.columns:
        merged = merged.rename(columns={"segment_prior": "segment_prev"})
        changers = changers.rename(columns={"segment_prior": "segment_prev"}) if "segment_prior" in changers.columns else changers

    # ── Transition matrix ─────────────────────────────────────────────────
    st.markdown("<div style='font-size:12px;font-weight:700;color:#6366f1;margin-bottom:6px;'>Transition Matrix (rows = Prior, cols = Current)</div>", unsafe_allow_html=True)
    matrix_data = {}
    for seg_p in SEG_ORDER:
        row_d = {}
        for seg_c in SEG_ORDER:
            mask = (merged.get("segment_prev", merged.get("segment_prior", pd.Series(dtype=str))) == seg_p) & (merged["segment_curr"] == seg_c)
            row_d[seg_c] = int(mask.sum())
        matrix_data[seg_p] = row_d
    matrix_df = pd.DataFrame(matrix_data).T.reindex(index=SEG_ORDER, columns=SEG_ORDER).fillna(0).astype(int)

    # Style the matrix as HTML
    header_cells = "".join(
        f"<th style='padding:6px 10px;font-size:11px;color:{SEG_COLORS.get(s,'#aaa')};text-align:center;'>{s}</th>"
        for s in SEG_ORDER
    )
    rows_html = ""
    for seg_p in SEG_ORDER:
        cells = ""
        for seg_c in SEG_ORDER:
            val = int(matrix_df.loc[seg_p, seg_c]) if seg_p in matrix_df.index else 0
            bg  = "#1a2a1a" if seg_p == seg_c and val > 0 else ("#2a1a1a" if val > 0 else "#0d0f1a")
            col = "#22c55e" if seg_p == seg_c and val > 0 else ("#fca5a5" if val > 0 else "#374151")
            cells += f"<td style='padding:6px 10px;text-align:center;background:{bg};color:{col};font-weight:{'700' if val > 0 else '400'};font-size:12px;border:1px solid #1e2235;'>{val}</td>"
        c = SEG_COLORS.get(seg_p, "#aaa")
        rows_html += f"<tr><td style='padding:6px 10px;font-size:11px;color:{c};font-weight:700;border:1px solid #1e2235;'>{seg_p}</td>{cells}</tr>"

    st.markdown(
        f"<table style='border-collapse:collapse;width:100%;margin-bottom:14px;'>"
        f"<thead><tr><th style='padding:6px 10px;font-size:11px;color:#64748b;'>Prior \ Current</th>{header_cells}</tr></thead>"
        f"<tbody>{rows_html}</tbody></table>",
        unsafe_allow_html=True,
    )

    # ── Movers list ───────────────────────────────────────────────────────
    seg_prev_col = "segment_prev" if "segment_prev" in merged.columns else "segment_prior"
    changers_clean = merged[merged["segment_curr"] != merged[seg_prev_col]].copy()
    n_movers = len(changers_clean)
    if n_movers == 0:
        st.success("No segment changes between the two periods. Pipeline is stable.")
        return

    with st.expander(f"🔀 {n_movers} Partner{'s' if n_movers > 1 else ''} Changed Segment — Click to Review", expanded=False):
        for _, row in changers_clean.iterrows():
            sp = str(row.get(seg_prev_col, "?"))
            sc = str(row.get("segment_curr", "?"))
            sp_c = SEG_COLORS.get(sp, "#aaa")
            sc_c = SEG_COLORS.get(sc, "#aaa")
            reason = _derive_change_reason(row)
            improved_flag = SEG_ORDER.index(sc) > SEG_ORDER.index(sp) if sp in SEG_ORDER and sc in SEG_ORDER else False
            arr_color = "#22c55e" if improved_flag else "#ef4444"
            arr = "↑" if improved_flag else "↓"
            st.markdown(
                f"<div style='background:#12141c;border:1px solid #1e2235;border-radius:8px;"
                f"padding:10px 14px;margin-bottom:6px;display:flex;align-items:center;gap:12px;'>"
                f"<div style='flex:2;font-size:13px;font-weight:600;color:#e2e8f0;'>{row.get('company_name','')}</div>"
                f"<div style='flex:1;font-size:12px;'>"
                f"<span style='color:{sp_c};font-weight:700;'>{sp}</span>"
                f"&nbsp;<span style='color:{arr_color};font-weight:700;'>{arr}</span>&nbsp;"
                f"<span style='color:{sc_c};font-weight:700;'>{sc}</span>"
                f"</div>"
                f"<div style='flex:3;font-size:11px;color:#94a3b8;'>{reason}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )


def _render_date_picker() -> tuple:
    """
    Render a premium calendar date-range picker with quick-select preset chips.
    Returns (start_date, end_date, date_label, using_custom_range).
    """
    # ── Inject chip CSS ────────────────────────────────────────────────────
    st.markdown("""
    <style>
    div[data-testid="column"] .stButton>button {
        border-radius: 20px !important;
        font-size: 12px !important;
        padding: 4px 14px !important;
        border: 1px solid #374151 !important;
        background: #1e2235 !important;
        color: #93c5fd !important;
        transition: all 0.2s ease !important;
    }
    div[data-testid="column"] .stButton>button:hover {
        background: #6366f1 !important;
        border-color: #6366f1 !important;
        color: #fff !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── State init ─────────────────────────────────────────────────────────
    today = date.today()
    if "kb_date_start" not in st.session_state:
        st.session_state["kb_date_start"] = today - timedelta(days=90)
    if "kb_date_end" not in st.session_state:
        st.session_state["kb_date_end"] = today
    if "kb_custom_active" not in st.session_state:
        st.session_state["kb_custom_active"] = False

    # ── Calendar header banner ─────────�

    # Row 3: Manual calendar pickers
    st.markdown(
        "<div style='font-size:11px;color:#64748b;margin-top:10px;margin-bottom:4px;'"
        ">🗓️ Custom Range</div>",
        unsafe_allow_html=True,
    )
    cal1, cal2, cal3 = st.columns([1.5, 1.5, 1])
    with cal1:
        new_start = st.date_input(
            "Start Date",
            value=st.session_state["kb_date_start"],
            max_value=today,
            key="kb_cal_start",
        )
    with cal2:
        new_end = st.date_input(
            "End Date",
            value=st.session_state["kb_date_end"],
            min_value=new_start,
            max_value=today,
            key="kb_cal_end",
        )
    with cal3:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("✅ Apply Range", key="kb_apply_range", use_container_width=True):
            st.session_state["kb_date_start"] = new_start
            st.session_state["kb_date_end"]   = new_end
            st.session_state["kb_custom_active"] = True

    # Resolve final dates
    sel_start = st.session_state["kb_date_start"]
    sel_end   = st.session_state["kb_date_end"]
    if sel_end < sel_start:
        sel_end = sel_start

    span_days  = max((sel_end - sel_start).days, 1)
    date_label = (
        f"{sel_start.strftime('%d %b %Y')} → {sel_end.strftime('%d %b %Y')}  ({span_days}d)"
    )
    using_custom = st.session_state.get("kb_custom_active", False)


    return sel_start, sel_end, date_label, span_days



def _render_filter_bar(df, cat_perf_df):
    """Render the inline main-page filter bar above the kanban board."""


    # ── Row 1: State/Area + Category + Sort ───────────────────────────────
    c1, c2, c3 = st.columns([2, 2, 1.5])

    with c1:
        all_states = sorted(df["state"].dropna().unique().tolist()) if "state" in df.columns else []
        sel_states = st.multiselect(
            "📍 Filter by Area / State",
            all_states,
            default=[],
            key="kb_states",
            help="Show only partners from selected states. Leave empty to show all.",
        )

    with c2:
        all_cats = []
        if not cat_perf_df.empty and "category" in cat_perf_df.columns:
            all_cats = sorted(cat_perf_df["category"].dropna().unique().tolist())
        sel_cats = st.multiselect(
            "🏷️ Filter by Product Category",
            all_cats,
            default=[],
            key="kb_cats",
            help="Show only partners who have purchased in the selected categories.",
        )

    with c3:
        sort_by = st.selectbox(
            "↕️ Sort by",
            ["Revenue (High→Low)", "Churn Risk (High→Low)", "At Risk (High→Low)", "Name (A→Z)"],
            key="kb_sort",
        )

    # ── Row 2: Revenue floor + Name search + Category toggle ──────────────
    c5, c6, c7 = st.columns([1.5, 2, 0.5])
    with c5:
        min_rev = st.number_input(
            "Min Revenue for Period (₹)",
            value=0, step=10000, key="kb_minrev",
        )
    with c6:
        search_query = st.text_input(
            "🔍 Search partner by name",
            placeholder="Type a company name...",
            key="kb_search",
        )
    with c7:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        show_cat_insight = st.toggle("Category View", key="kb_cat_view", value=False)

    return sel_states, sel_cats, sort_by, min_rev, search_query, show_cat_insight


def _render_top_at_risk_alert(df: pd.DataFrame, rev_mult: float):
    """
    Show a prominent warning if any top-tier partner (top-20% recent revenue)
    has landed in At Risk or Critical. These are high-value accounts that
    require immediate human review — not just algorithmic flagging.
    """
    if df.empty or "recent_90_revenue" not in df.columns:
        return
    rev_p80 = float((df["recent_90_revenue"] * rev_mult).quantile(0.80))
    dangerous = df[
        ((df["recent_90_revenue"] * rev_mult) >= rev_p80)
        & (df["health_segment"].isin({"At Risk", "Critical"}))
    ].copy()
    if dangerous.empty:
        return

    dangerous = dangerous.sort_values("recent_90_revenue", ascending=False)
    st.markdown(
        f"""
        <div style='background:linear-gradient(135deg,#1c0a0a,#2a0e0e);
             border:1px solid #7f1d1d;border-radius:12px;
             padding:14px 20px;margin-bottom:18px;'>
          <div style='font-size:13px;font-weight:800;color:#fca5a5;margin-bottom:6px;'>
            🚨 Elite Partners Needing Urgent Attention ({len(dangerous)} account{"s" if len(dangerous)>1 else ""})
          </div>
          <div style='font-size:12px;color:#f87171;'>
            These top-20% revenue partners are currently flagged At Risk or Critical.
            Review immediately — revenue at stake is significant.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    cols = st.columns(min(len(dangerous), 4))
    for idx, (_, row) in enumerate(dangerous.head(4).iterrows()):
        name      = str(row.get("company_name", "Unknown"))
        rev       = _fmt_inr(float(row.get("recent_90_revenue", 0) or 0) * rev_mult)
        seg       = str(row.get("health_segment", "At Risk"))
        seg_color = "#f59e0b" if seg == "At Risk" else "#ef4444"
        churn     = float(row.get("churn_probability", 0) or 0)
        drop      = float(row.get("revenue_drop_pct", 0) or 0)
        with cols[idx]:
            st.markdown(
                f"""<div style='background:#1f0d0d;border:1px solid {seg_color}55;
                     border-left:4px solid {seg_color};border-radius:10px;
                     padding:12px 14px;'>
                  <div style='font-size:13px;font-weight:700;color:#fef2f2;
                       margin-bottom:6px;line-height:1.4;'>{name}</div>
                  <div style='font-size:20px;font-weight:800;color:#fca5a5;
                       margin-bottom:4px;'>{rev}</div>
                  <div style='font-size:11px;'>
                    <span style='color:{seg_color};font-weight:700;'>● {seg}</span>
                    &nbsp;|&nbsp;
                    <span style='color:#94a3b8;'>Churn {churn:.0%}</span>
                  </div>
                  {'<div style="font-size:11px;color:#ef4444;margin-top:4px;">📉 Rev ↓' + f'{drop:.0f}%</div>' if drop > 0 else ''}
                </div>""",
                unsafe_allow_html=True,
            )


def _render_category_insight(cat_perf_df, kanban_df):
    """
    Renders a category performance breakdown table:
    For each category → best performing distributor and worst performing distributor.
    """
    if cat_perf_df.empty or "category" not in cat_perf_df.columns:
        st.info("No category spend data available.")
        return

    st.markdown("""
    <div style='background:#12141c;border:1px solid #1e2235;border-radius:12px;
         padding:14px 20px;margin-bottom:18px;'>
      <div style='font-size:13px;font-weight:700;color:#6366f1;margin-bottom:4px;'>
        📊 Category Performance — Distributor Breakdown
      </div>
      <div style='font-size:12px;color:#64748b;'>
        Which distributors are performing best and worst in each product category
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Merge health segment info
    seg_map = {}
    if not kanban_df.empty and "health_segment" in kanban_df.columns:
        seg_map = kanban_df.set_index("company_name")["health_segment"].to_dict()

    # Aggregate: sum spend per company per category
    cat_agg = (
        cat_perf_df.groupby(["category", "company_name"])["total_spend"]
        .sum()
        .reset_index()
    )
    cat_agg["health"] = cat_agg["company_name"].map(seg_map).fillna("—")

    categories = sorted(cat_agg["category"].unique().tolist())

    for cat in categories:
        cat_slice = cat_agg[cat_agg["category"] == cat].sort_values("total_spend", ascending=False)
        if cat_slice.empty:
            continue

        best  = cat_slice.head(3)
        worst = cat_slice.tail(3).sort_values("total_spend")

        with st.expander(f"🏷️ {cat}  •  {len(cat_slice)} distributors  •  Total: {_fmt_inr(cat_slice['total_spend'].sum())}"):
            col_b, col_w = st.columns(2)

            with col_b:
                st.markdown("<div style='color:#22c55e;font-weight:700;font-size:12px;margin-bottom:6px;'>🥇 Top Performers</div>", unsafe_allow_html=True)
                for _, r in best.iterrows():
                    seg = r["health"]
                    seg_color = {"Champion": "#22c55e", "Emerging": "#06b6d4", "Healthy": "#3b82f6",
                                 "At Risk": "#f59e0b", "Critical": "#ef4444"}.get(seg, "#aaa")
                    st.markdown(
                        f"<div style='background:#0d1f0d;border-left:3px solid #22c55e;"
                        f"border-radius:6px;padding:8px 12px;margin-bottom:5px;'>"
                        f"<span style='font-weight:600;color:#e2e8f0;font-size:13px;'>{r['company_name']}</span>"
                        f"<span style='float:right;color:#22c55e;font-weight:700;'>{_fmt_inr(r['total_spend'])}</span><br/>"
                        f"<span style='font-size:11px;color:{seg_color};'>● {seg}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

            with col_w:
                st.markdown("<div style='color:#ef4444;font-weight:700;font-size:12px;margin-bottom:6px;'>⚠️ Lowest Performers</div>", unsafe_allow_html=True)
                for _, r in worst.iterrows():
                    seg = r["health"]
                    seg_color = {"Champion": "#22c55e", "Emerging": "#06b6d4", "Healthy": "#3b82f6",
                                 "At Risk": "#f59e0b", "Critical": "#ef4444"}.get(seg, "#aaa")
                    st.markdown(
                        f"<div style='background:#1f0d0d;border-left:3px solid #ef4444;"
                        f"border-radius:6px;padding:8px 12px;margin-bottom:5px;'>"
                        f"<span style='font-weight:600;color:#e2e8f0;font-size:13px;'>{r['company_name']}</span>"
                        f"<span style='float:right;color:#ef4444;font-weight:700;'>{_fmt_inr(r['total_spend'])}</span><br/>"
                        f"<span style='font-size:11px;color:{seg_color};'>● {seg}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )


# ── Excel Export ───────────────────────────────────────────────────────────
def _export_excel(df: pd.DataFrame, rev_mult: float, period_label: str) -> bytes:
    """
    Build a styled Excel workbook with:
      • Sheet 1 – Pipeline Summary (metrics + per-lane breakdown)
      • Sheet 2–6 – One sheet per Kanban lane with full partner data
    Returns the workbook as bytes ready for st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    LANE_COLORS = {
        "Champion": ("1E3A24", "22C55E"),
        "Emerging": ("0E2F38", "06B6D4"),
        "Healthy":  ("0F1F3D", "3B82F6"),
        "At Risk":  ("3A2E10", "F59E0B"),
        "Critical": ("3A1212", "EF4444"),
    }
    HEADER_FILL  = PatternFill("solid", fgColor="1A1C2E")
    HEADER_FONT  = Font(bold=True, color="E2E8F0", size=11, name="Calibri")
    SUB_FILL     = PatternFill("solid", fgColor="12141C")
    SUB_FONT     = Font(bold=True, color="93C5FD", size=10, name="Calibri")
    CELL_FONT    = Font(color="CBD5E1", size=10, name="Calibri")
    TITLE_FONT   = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    thin_side    = Side(style="thin", color="374151")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    def _money(val):
        try:
            return round(float(val), 2)
        except Exception:
            return 0.0

    def _pct(val):
        try:
            return round(float(val) * 100, 1)
        except Exception:
            return 0.0

    def _auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                min(max(max_len + 2, min_w), max_w)
            )

    def _style_header_row(ws, row_idx, columns, fill_hex, font_hex="FFFFFF"):
        hdr_fill = PatternFill("solid", fgColor=fill_hex)
        hdr_font = Font(bold=True, color=font_hex, size=10, name="Calibri")
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def _write_lane_sheet(wb, lane, lane_df, rev_mult):
        seg = lane["label"].split(" ", 1)[1]  # strip emoji
        seg_key = list(lane["segments"])[0]
        bg_hex, accent_hex = LANE_COLORS.get(seg_key, ("1A1C2E", "6366F1"))
        ws = wb.create_sheet(title=seg_key[:31])

        # Title row
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"{lane['label']} — {period_label} View"
        title_cell.fill = PatternFill("solid", fgColor=bg_hex)
        title_cell.font = Font(bold=True, color=accent_hex, size=14, name="Calibri")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Header
        headers = [
            "Partner Name", "State", f"{period_label} Revenue (₹)",
            "Revenue Trend (%)", "Churn Probability (%)",
            "Credit Risk", "Revenue at Risk (₹)", "Cluster", "Last Order"
        ]
        for col_i, h in enumerate(headers, start=1):
            ws.cell(row=2, column=col_i, value=h)
        _style_header_row(ws, 2, headers, bg_hex, accent_hex)

        # Data rows
        for r_i, (_, row) in enumerate(lane_df.iterrows(), start=3):
            fill_color = "12141C" if r_i % 2 == 0 else "0D0F1A"
            row_fill = PatternFill("solid", fgColor=fill_color)
            values = [
                str(row.get("company_name", "")),
                str(row.get("state", "") or ""),
                _money(row.get("recent_90_revenue", 0) or 0) * rev_mult,
                _pct(-(row.get("revenue_drop_pct", 0) or 0)),  # positive=growth
                _pct(row.get("churn_probability", 0) or 0),
                str(row.get("credit_risk_band", "") or ""),
                _money(row.get("revenue_at_risk", 0) or 0) * rev_mult,
                str(row.get("cluster_label", "") or ""),
                _days_ago(row.get("recency_days", None)),
            ]
            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=r_i, column=col_i, value=val)
                cell.fill = row_fill
                cell.font = CELL_FONT
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A3"
        _auto_width(ws)

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    # Title
    ws_sum.merge_cells("A1:F1")
    t = ws_sum["A1"]
    t.value = f"Revenue Pipeline Tracker — {period_label} Export"
    t.fill = PatternFill("solid", fgColor="0F1117")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 36

    # Generated date
    ws_sum.merge_cells("A2:F2")
    d = ws_sum["A2"]
    d.value = f"Generated: {date.today().strftime('%d %B %Y')}"
    d.fill = PatternFill("solid", fgColor="12141C")
    d.font = Font(italic=True, color="64748B", size=10, name="Calibri")
    d.alignment = Alignment(horizontal="center")

    # KPI section header
    ws_sum.merge_cells("A4:F4")
    kh = ws_sum["A4"]
    kh.value = "Pipeline Summary Metrics"
    kh.fill = HEADER_FILL
    kh.font = HEADER_FONT
    kh.alignment = Alignment(horizontal="center")

    total_rev    = float(df["recent_90_revenue"].sum()) * rev_mult
    total_risk   = float(df["revenue_at_risk"].sum()) * rev_mult
    high_churn   = int((df["churn_probability"] >= 0.65).sum())
    critical_cnt = int((df.get("health_segment", pd.Series(dtype=str)) == "Critical").sum())

    kpis = [
        ("Total Partners", len(df)),
        (f"{period_label} Pipeline Value (₹)", round(total_rev, 2)),
        ("Revenue at Risk (₹)", round(total_risk, 2)),
        ("High Churn Risk Partners", high_churn),
        ("Critical Accounts", critical_cnt),
    ]
    kpi_headers = ["Metric", "Value"]
    for col_i, h in enumerate(kpi_headers, start=1):
        ws_sum.cell(row=5, column=col_i, value=h)
    _style_header_row(ws_sum, 5, kpi_headers, "1E2235", "93C5FD")

    for r_i, (label, val) in enumerate(kpis, start=6):
        fill_color = "12141C" if r_i % 2 == 0 else "0D0F1A"
        ws_sum.cell(row=r_i, column=1, value=label).fill = PatternFill("solid", fgColor=fill_color)
        ws_sum.cell(row=r_i, column=1).font = CELL_FONT
        ws_sum.cell(row=r_i, column=1).border = thin_border
        ws_sum.cell(row=r_i, column=2, value=val).fill = PatternFill("solid", fgColor=fill_color)
        ws_sum.cell(row=r_i, column=2).font = Font(bold=True, color="F8FAFC", size=10, name="Calibri")
        ws_sum.cell(row=r_i, column=2).border = thin_border

    # Lane breakdown
    row_start = len(kpis) + 8
    ws_sum.merge_cells(f"A{row_start}:F{row_start}")
    bh = ws_sum.cell(row=row_start, column=1, value="Lane Breakdown")
    bh.fill = HEADER_FILL
    bh.font = HEADER_FONT
    bh.alignment = Alignment(horizontal="center")

    lane_hdr_row = row_start + 1
    lane_headers = ["Lane", "Partners", f"Revenue (₹)", "Revenue at Risk (₹)", "Avg Churn (%)"]
    for col_i, h in enumerate(lane_headers, start=1):
        ws_sum.cell(row=lane_hdr_row, column=col_i, value=h)
    _style_header_row(ws_sum, lane_hdr_row, lane_headers, "1E2235", "93C5FD")

    for r_i, lane in enumerate(LANES, start=lane_hdr_row + 1):
        seg_key = list(lane["segments"])[0]
        lane_mask = df["health_segment"].isin(lane["segments"]) if "health_segment" in df.columns else pd.Series(False, index=df.index)
        lane_df = df[lane_mask]
        bg_hex, accent_hex = LANE_COLORS.get(seg_key, ("1A1C2E", "6366F1"))
        vals = [
            lane["label"],
            len(lane_df),
            round(float(lane_df["recent_90_revenue"].sum()) * rev_mult, 2),
            round(float(lane_df["revenue_at_risk"].sum()) * rev_mult, 2),
            round(float(lane_df["churn_probability"].mean()) * 100, 1) if not lane_df.empty else 0.0,
        ]
        fill = PatternFill("solid", fgColor=bg_hex)
        for col_i, val in enumerate(vals, start=1):
            cell = ws_sum.cell(row=r_i, column=col_i, value=val)
            cell.fill = fill
            cell.font = Font(bold=(col_i == 1), color=accent_hex if col_i == 1 else "CBD5E1",
                             size=10, name="Calibri")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    _auto_width(ws_sum)

    # Lane sheets
    for lane in LANES:
        seg_key = list(lane["segments"])[0]
        lane_mask = df["health_segment"].isin(lane["segments"]) if "health_segment" in df.columns else pd.Series(False, index=df.index)
        lane_df = df[lane_mask]
        _write_lane_sheet(wb, lane, lane_df, rev_mult)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF Export ────────────────────────────────────────────────────────────


def _kb_pdf_safe(text: str, maxlen: int = 0) -> str:
    """
    Sanitize text for FPDF2 (Helvetica core font = latin-1 only).
    - Replaces common Unicode arrows/symbols with ASCII equivalents.
    - Strips ALL remaining non-latin-1 characters (including emojis,
      Devanagari, RTL marks etc.) so FPDF never throws UnicodeEncodeError.
    - Optionally truncates to maxlen chars (0 = no limit).
    """
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    repl = {
        '\u2192': '->', '\u2190': '<-', '\u2191': '^', '\u2193': 'v',
        '\u20b9': 'Rs', '\u2022': '-', '\u2013': '-', '\u2014': '-',
        '\u2019': "'", '\u2018': "'", '\u201c': '"', '\u201d': '"',
        '\u2026': '...', '\u00b7': '-', '\u00d7': 'x',
        '\u001a': '-', '\u2248': '~', '\u2260': '!=',
        # Common emoji replacements (lane labels)
        '\U0001f3c6': '[Champ]',  # 🏆
        '\U0001f680': '[Rise]',   # 🚀
        '\u2705': '[OK]',         # ✅
        '\u26a0': '[Warn]',       # ⚠
        '\ufe0f': '',             # variation selector
        '\U0001f534': '[Crit]',   # 🔴
        '\U0001f4e6': '[Pkg]',    # 📦
        '\U0001f4cd': '[Pin]',    # 📍
        '\U0001f3f7': '[Tag]',    # 🏷
        '\u2b50': '[Star]',       # ⭐
        '\U0001f4c5': '[Cal]',    # 📅
        '\U0001f4c8': '[Chart]',  # 📈
        '\u26a1': '[!]',          # ⚡
        '\U0001f4c9': '[v]',      # 📉
        '\U0001f6a8': '[SOS]',    # 🚨
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    # Final pass: drop anything still outside latin-1
    text = text.encode('latin-1', errors='ignore').decode('latin-1')
    if maxlen and len(text) > maxlen:
        text = text[:maxlen - 1] + '.'
    return text

def _export_pdf(df: pd.DataFrame, rev_mult: float, period_label: str) -> bytes:
    """
    Generate a formatted PDF report:
      • Cover header with generation date
      • Pipeline KPI summary table
      • Per-lane section with partner table
    Returns bytes ready for st.download_button.
    Robust against UnicodeEncodeError — all strings are sanitised via _kb_pdf_safe()
    before being passed to FPDF cells.
    """
    try:
        from fpdf import FPDF
    except ImportError as exc:
        raise ImportError(
            "fpdf2 is required for PDF export. Install it with: pip install fpdf2"
        ) from exc

    LANE_RGB = {
        "Champion": (34, 197, 94),
        "Emerging": (6, 182, 212),
        "Healthy":  (59, 130, 246),
        "At Risk":  (245, 158, 11),
        "Critical": (239, 68, 68),
    }
    BG   = (13, 15, 23)
    CARD = (26, 28, 35)
    TEXT = (203, 213, 225)
    DIM  = (100, 116, 139)
    WHT  = (255, 255, 255)
    ACC  = (99, 102, 241)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(12, 12, 12)

    # ── Cover page ─────────────────────────────────────────────────────────
    pdf.add_page()
    # Background
    pdf.set_fill_color(*BG)
    pdf.rect(0, 0, 297, 210, "F")

    # Title bar
    pdf.set_fill_color(*ACC)
    pdf.rect(0, 0, 297, 28, "F")
    pdf.set_text_color(*WHT)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_xy(12, 6)
    pdf.cell(0, 12, "Revenue Pipeline Tracker", ln=False)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_xy(12, 18)
    pdf.cell(0, 8, _kb_pdf_safe(f"{period_label} Export  |  Generated: {date.today().strftime('%d %B %Y')}"))

    # KPI summary
    total_rev    = float(df["recent_90_revenue"].sum()) * rev_mult
    total_risk   = float(df["revenue_at_risk"].sum()) * rev_mult
    high_churn   = int((df["churn_probability"] >= 0.65).sum())
    critical_cnt = int((df.get("health_segment", pd.Series(dtype=str)) == "Critical").sum())

    def _fmt_pdf(val):
        try:
            v = float(val)
        except Exception:
            return "0"
        if v >= 1_00_00_000: return f"Rs {v/1_00_00_000:.1f} Cr"
        if v >= 1_00_000:    return f"Rs {v/1_00_000:.1f} L"
        if v >= 1_000:       return f"Rs {v/1_000:.0f} K"
        return f"Rs {v:.0f}"

    kpis = [
        ("Partners in Pipeline",    str(len(df))),
        (_kb_pdf_safe(f"{period_label} Value"),   _fmt_pdf(total_rev)),
        ("Revenue at Risk",         _fmt_pdf(total_risk)),
        ("High Churn Risk",         str(high_churn)),
        ("Critical Accounts",       str(critical_cnt)),
    ]
    card_w = (297 - 24 - 16) / len(kpis)
    pdf.set_y(36)
    for i, (label, val) in enumerate(kpis):
        x = 12 + i * (card_w + 4)
        pdf.set_fill_color(*CARD)
        pdf.set_draw_color(*DIM)
        pdf.set_line_width(0.3)
        pdf.rect(x, 36, card_w, 28, "FD")
        pdf.set_text_color(*DIM)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_xy(x + 2, 39)
        pdf.cell(card_w - 4, 5, label, align="C")
        pdf.set_text_color(*WHT)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_xy(x + 2, 46)
        pdf.cell(card_w - 4, 10, val, align="C")

    # Lane summary table
    pdf.set_y(72)
    pdf.set_text_color(*ACC)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Pipeline Lane Summary", ln=True)

    col_ws = [60, 22, 48, 48, 35]
    headers = ["Lane", "Count", f"Revenue (Approx)", "At Risk (Approx)", "Avg Churn"]
    pdf.set_fill_color(*CARD)
    pdf.set_text_color(*WHT)
    pdf.set_font("Helvetica", "B", 9)
    for w, h in zip(col_ws, headers):
        pdf.cell(w, 7, h, border=1, fill=True, align="C")
    pdf.ln()

    for lane in LANES:
        seg_key = list(lane["segments"])[0]
        r, g, b = LANE_RGB.get(seg_key, (99, 102, 241))
        lane_mask = df["health_segment"].isin(lane["segments"]) if "health_segment" in df.columns else pd.Series(False, index=df.index)
        lane_df = df[lane_mask]
        lane_rev  = float(lane_df["recent_90_revenue"].sum()) * rev_mult
        lane_risk = float(lane_df["revenue_at_risk"].sum()) * rev_mult
        lane_churn = float(lane_df["churn_probability"].mean()) * 100 if not lane_df.empty else 0.0

        pdf.set_fill_color(int(r*0.15), int(g*0.15), int(b*0.15))
        pdf.set_text_color(r, g, b)
        pdf.set_font("Helvetica", "B", 9)
        _lane_label_safe = _kb_pdf_safe(lane["label"], maxlen=30)
        pdf.cell(col_ws[0], 7, _lane_label_safe, border=1, fill=True)
        pdf.set_text_color(*TEXT)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(col_ws[1], 7, _kb_pdf_safe(str(len(lane_df))), border=1, align="C")
        pdf.cell(col_ws[2], 7, _kb_pdf_safe(_fmt_pdf(lane_rev)), border=1, align="R")
        pdf.cell(col_ws[3], 7, _kb_pdf_safe(_fmt_pdf(lane_risk)), border=1, align="R")
        pdf.cell(col_ws[4], 7, _kb_pdf_safe(f"{lane_churn:.1f}%"), border=1, align="C")
        pdf.ln()

    # ── Per-lane detail pages ──────────────────────────────────────────────
    partner_cols = [
        ("company_name",       "Partner",         60),
        ("state",              "State",            22),
        ("recent_90_revenue",  "Revenue (Rs)",    42),
        ("churn_probability",  "Churn %",          22),
        ("revenue_at_risk",    "At Risk (Rs)",     42),
        ("credit_risk_band",   "Credit",           22),
        ("recency_days",       "Last Order",       26),
        ("cluster_label",      "Cluster",          36),
    ]

    for lane in LANES:
        seg_key = list(lane["segments"])[0]
        r, g, b = LANE_RGB.get(seg_key, (99, 102, 241))
        lane_mask = df["health_segment"].isin(lane["segments"]) if "health_segment" in df.columns else pd.Series(False, index=df.index)
        lane_df = df[lane_mask]
        if lane_df.empty:
            continue

        pdf.add_page()
        pdf.set_fill_color(*BG)
        pdf.rect(0, 0, 297, 210, "F")

        # Lane title bar
        pdf.set_fill_color(r, g, b)
        pdf.rect(0, 0, 297, 14, "F")
        pdf.set_text_color(*WHT)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_xy(12, 2)
        pdf.cell(0, 10, _kb_pdf_safe(f"{lane['label']} Partners  ({len(lane_df)} accounts)  |  {period_label}", maxlen=80))

        # Table header
        pdf.set_y(18)
        pdf.set_fill_color(*CARD)
        pdf.set_text_color(r, g, b)
        pdf.set_font("Helvetica", "B", 8)
        for (_, hdr, w) in partner_cols:
            pdf.cell(w, 6, hdr, border=1, fill=True, align="C")
        pdf.ln()

        # Rows (cap at 100 per lane for pdf size)
        pdf.set_font("Helvetica", "", 8)
        for row_i, (_, row) in enumerate(lane_df.head(100).iterrows()):
            fill_toggle = row_i % 2 == 0
            pdf.set_fill_color(*(CARD if fill_toggle else BG))
            for (col_key, _, w) in partner_cols:
                raw = row.get(col_key, "")
                if col_key == "recent_90_revenue":
                    val_str = _kb_pdf_safe(_fmt_pdf(float(raw or 0) * rev_mult))
                elif col_key == "revenue_at_risk":
                    val_str = _kb_pdf_safe(_fmt_pdf(float(raw or 0) * rev_mult))
                elif col_key == "churn_probability":
                    try:
                        val_str = f"{float(raw)*100:.0f}%"
                    except Exception:
                        val_str = "-"
                elif col_key == "recency_days":
                    val_str = _kb_pdf_safe(_days_ago(raw))
                else:
                    val_str = _kb_pdf_safe(str(raw or "-"), maxlen=25)
                pdf.set_text_color(*TEXT)
                pdf.cell(w, 5.5, val_str, border=1, fill=True, align="L")
            pdf.ln()

        if len(lane_df) > 100:
            pdf.set_text_color(*DIM)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 6, f"  Showing 100 of {len(lane_df)} partners. Use Excel export for full data.", ln=True)

    return bytes(pdf.output())


# ── Export button row ─────────────────────────────────────────────────────
def _render_export_buttons(df: pd.DataFrame, rev_mult: float, period_label: str):
    """Render Excel and PDF export download buttons above the Kanban board."""
    st.markdown(
        """
        <div style='background:#12141c;border:1px solid #1e2235;border-radius:12px;
             padding:14px 20px;margin-bottom:18px;display:flex;align-items:center;gap:12px;'>
          <span style='font-size:18px;'>📤</span>
          <span style='font-size:13px;font-weight:700;color:#6366f1;'>Export Pipeline Report</span>
          <span style='font-size:12px;color:#64748b;'>Download the current filtered view as Excel or PDF</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    col_xl, col_pdf, col_spacer = st.columns([1, 1, 4])

    with col_xl:
        try:
            excel_bytes = _export_excel(df, rev_mult, period_label)
            filename_xl = f"pipeline_{period_label.lower()}_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_bytes,
                file_name=filename_xl,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="kb_export_excel",
                help="Download a styled Excel workbook with one sheet per pipeline lane",
            )
        except Exception as e:
            st.error(f"Excel export failed: {e}")

    with col_pdf:
        try:
            pdf_bytes = _export_pdf(df, rev_mult, period_label)
            # _export_pdf already returns bytes — no double-conversion needed
            filename_pdf = f"pipeline_{date.today().strftime('%Y%m%d')}.pdf"
            st.download_button(
                label="⬇️ Download PDF",
                data=pdf_bytes,
                file_name=filename_pdf,
                mime="application/pdf",
                use_container_width=True,
                key="kb_export_pdf",
                help="Download a formatted PDF report with pipeline summary and per-lane partner tables",
            )
        except ImportError:
            st.warning("PDF export requires fpdf2. Run: `pip install fpdf2`")
        except Exception as e:
            st.error(f"PDF export failed: {e}")


def render(ai):

    page_header(
        title="Revenue Pipeline Tracker",
        subtitle="Monitor partner health across every stage — Champion, Emerging, Healthy, At Risk, and Critical.",
        icon="📊",
        accent_color="#6366f1",
    )
    skel = st.empty()
    with skel.container():
        skeleton_loader(n_metric_cards=5, n_rows=2, label="Loading pipeline data...")

    # Load clustering (fast, cached) and optionally churn/credit
    ai.ensure_clustering()
    if getattr(ai, "enable_realtime_partner_scoring", False):
        try:
            ai.ensure_churn_forecast()
            ai.ensure_credit_risk()
        except Exception:
            pass

    skel.empty()

    pf = getattr(ai, "df_partner_features", None)
    if pf is None or pf.empty:
        st.warning("Partner features not available. Run the clustering engine first.")
        return

    # Use cached, lightweight slice
    # Fix A: pass a content-hash scalar so @st.cache_data keys on data, not object id
    _pf_hash = str(hash(tuple(pf.index.tolist())))[:16]
    df = _build_kanban_df(pf, _pf_hash)

    # Category performance data from df_recent_group_spend
    group_spend = getattr(ai, "df_recent_group_spend", None)
    cat_perf_df = _build_category_performance(group_spend)

    # ── Date range picker (calendar UI) ─────────────────────────────────────
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    sel_start, sel_end, date_label, span_days = _render_date_picker()

    # ── Main-page filter bar ──────────────────────────────────────────────
    sel_states, sel_cats, sort_by, min_rev, search_query, show_cat_insight = (
        _render_filter_bar(df, cat_perf_df)
    )

    # ── Fetch actual revenue for the selected custom date window ────────────
    # rev_mult is used for display scaling of at-risk / concentration metrics.
    # Since we now have REAL revenue for the period, mult = 1 (no scaling needed).
    # We still set it proportionally so "Revenue at Risk" stays period-consistent.
    rev_mult    = 1.0          # actual revenue is already the exact period amount
    period_label = date_label  # e.g. "01 Dec 2025 → 31 Dec 2025 (31d)"

    engine = getattr(ai, "engine", None)

    # Fix B: session_state revenue cache — avoids repeat DB hits on same dates
    _rev_key = f"rev_{sel_start}_{sel_end}"
    if _rev_key not in st.session_state:
        st.session_state[_rev_key] = _fetch_custom_period_revenue(engine, sel_start, sel_end)
    custom_rev_df = st.session_state[_rev_key]

    if not custom_rev_df.empty and "company_name" in custom_rev_df.columns:
        # Merge actual period revenue onto the base kanban df
        # (keeps all metadata: state, cluster, churn, credit etc.)
        df = df.merge(
            custom_rev_df[["company_name", "period_revenue", "period_txns", "last_purchase_date"]],
            on="company_name",
            how="left",
        )
        df["period_revenue"] = pd.to_numeric(df["period_revenue"], errors="coerce").fillna(0.0)
        # Override recent_90_revenue with the real period revenue so all
        # downstream logic (segments, metrics, export) uses the actual figure
        df["recent_90_revenue"] = df["period_revenue"]

        # Recompute recency_days relative to sel_end
        if "last_purchase_date" in df.columns:
            _lp = pd.to_datetime(df["last_purchase_date"], errors="coerce")
            _end_ts = pd.Timestamp(sel_end)
            df["recency_days"] = (_end_ts - _lp).dt.days.fillna(9999)

        # Drop the helper column
        df = df.drop(columns=[c for c in ["period_revenue", "period_txns", "last_purchase_date"]
                              if c in df.columns], errors="ignore")

        # Recompute revenue_at_risk from updated churn * period revenue
        _MIN_FLOOR = 5_000 * max(span_days / 30, 1)
        _base = df["churn_probability"] * df["recent_90_revenue"]
        _floor_mask = (df["recent_90_revenue"] == 0) & (df["churn_probability"] >= 0.5)
        df["revenue_at_risk"] = _base.where(~_floor_mask, df["churn_probability"] * _MIN_FLOOR)
    else:
        # No DB connection or no data for the period — scale 90d actuals
        _scale = span_days / 90.0
        rev_mult = _scale

    # ── Apply state filter ────────────────────────────────────────────────
    if sel_states:
        df = df[df["state"].isin(sel_states)]

    # ── Apply category filter ─────────────────────────────────────────────
    if sel_cats and not cat_perf_df.empty and "category" in cat_perf_df.columns:
        cat_companies = set(
            cat_perf_df[cat_perf_df["category"].isin(sel_cats)]["company_name"].unique()
        )
        df = df[df["company_name"].isin(cat_companies)]

    # ── Apply revenue floor (compare against scaled period revenue) ──────
    if min_rev > 0:
        df = df[(df["recent_90_revenue"] * rev_mult) >= min_rev]

    # ── Apply name search ─────────────────────────────────────────────────
    if search_query.strip():
        df = df[df["company_name"].str.contains(search_query.strip(), case=False, na=False)]

    # ── Period-aware health re-segmentation ──────────────────────────────
    # Fix C: lazy recompute — key includes ALL filter state to avoid stale cache
    _filter_sig = f"{sorted(sel_states)}_{sorted(sel_cats)}_{min_rev}_{search_query.strip()}"
    _seg_key = f"seg_{_pf_hash}_{sel_start}_{sel_end}_{hash(_filter_sig) % 999999}"
    if _seg_key not in st.session_state:
        st.session_state[_seg_key] = _recompute_segments_for_period(df, rev_mult, period="Quarterly")
    df = st.session_state[_seg_key]

    # ── Sort ──────────────────────────────────────────────────────────────
    if sort_by == "Revenue (High→Low)":
        df = df.sort_values("recent_90_revenue", ascending=False)
    elif sort_by == "Churn Risk (High→Low)":
        df = df.sort_values("churn_probability", ascending=False)
    elif sort_by == "At Risk (High→Low)":
        df = df.sort_values("revenue_at_risk", ascending=False)
    else:
        df = df.sort_values("company_name")

    st.markdown("---")

    # ── Category insight view (toggle) ────────────────────────────────────
    if show_cat_insight:
        _render_category_insight(cat_perf_df, df)
        st.markdown("---")

    # ── Summary bar — 6 metrics ───────────────────────────────────────────
    total_partners   = len(df)
    high_churn       = int((df["churn_probability"] >= 0.65).sum())
    critical_cnt     = int((df.get("health_segment", pd.Series(dtype=str)) == "Critical").sum())
    total_revenue    = float(df["recent_90_revenue"].sum()) * rev_mult
    total_at_risk    = float(df["revenue_at_risk"].sum()) * rev_mult
    # period_label is already set from date_label / date picker above

    # Revenue concentration: top-10 partners' share of total
    _top10_rev = float(
        df.nlargest(10, "recent_90_revenue")["recent_90_revenue"].sum() * rev_mult
    ) if not df.empty else 0.0
    _conc_pct  = (_top10_rev / total_revenue * 100) if total_revenue > 0 else 0.0

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Partners in Pipeline", total_partners)
    m2.metric("High Churn Risk", high_churn,
              delta=f"⚠️ {high_churn/max(total_partners,1)*100:.0f}% of Pipeline",
              delta_color="inverse")
    m3.metric("Critical Accounts", critical_cnt, delta_color="inverse")
    m4.metric(f"{period_label} Pipeline Value", _fmt_inr(total_revenue))
    m5.metric("Revenue at Risk ⓘ", _fmt_inr(total_at_risk),
              delta="⚠️ AI Risk + Critical", delta_color="inverse")
    m6.metric("Top-10 Concentration", f"{_conc_pct:.0f}%",
              delta=f"{_fmt_inr(_top10_rev)} in top 10",
              delta_color="off",
              help="Revenue share from your top-10 partners. High concentration = key account risk.")

    # Active filter info chips
    active_filters = []
    if sel_states:
        active_filters.append(f"📍 {', '.join(sel_states)}")
    if sel_cats:
        active_filters.append(f"🏷️ {', '.join(sel_cats)}")
    active_filters.append(f"📅 {period_label}")
    if min_rev > 0:
        active_filters.append(f"₹ ≥ {_fmt_inr(min_rev)}")
    if search_query.strip():
        active_filters.append(f"🔍 \"{search_query.strip()}\"")

    if active_filters:
        chips_html = "".join(
            f"<span style='background:#1e2235;border:1px solid #374151;color:#93c5fd;"
            f"border-radius:20px;padding:3px 10px;font-size:11px;margin-right:6px;'>{f}</span>"
            for f in active_filters
        )
        st.markdown(
            f"<div style='margin-top:6px;margin-bottom:2px;'>"
            f"<span style='color:#64748b;font-size:11px;margin-right:8px;'>Active filters:</span>"
            f"{chips_html}</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")

    # ── Segment Migration Panel (Before vs After) — always visible ——————
    # Show when span_days > 0 and engine is available
    _base_df_for_migration = _build_kanban_df(pf, _pf_hash).copy()
    _render_segment_migration(
        df_curr=df,
        base_df=_base_df_for_migration,
        engine=engine,
        sel_start=sel_start,
        sel_end=sel_end,
        span_days=span_days,
        rev_mult=rev_mult,
    )

    st.markdown("---")

    # ── Top-tier partners at risk: early-warning alert ────────────────────
    _render_top_at_risk_alert(df, rev_mult)

    # ── Export buttons ────────────────────────────────────────────────────
    _render_export_buttons(df, rev_mult, period_label)

    # ── Kanban Board ──────────────────────────────────────────────────────
    cols = st.columns(len(LANES))

    for idx, lane in enumerate(LANES):
        col = cols[idx]
        mask = (
            df["health_segment"].isin(lane["segments"])
            if "health_segment" in df.columns
            else pd.Series(False, index=df.index)
        )
        lane_df = df[mask]
        lane_count = len(lane_df)
        lane_rev   = float(lane_df["recent_90_revenue"].sum()) * rev_mult
        lane_risk  = float(lane_df["revenue_at_risk"].sum()) * rev_mult

        with col:
            # Lane header with playbook tooltip
            playbook = lane.get("playbook", "")
            tier_badge = lane.get("tier", "")
            st.markdown(
                f"""<div style="background:#1a1c23;padding:12px;border-top:4px solid {lane['color']};border-radius:8px;margin-bottom:12px;">
                    <h4 style="margin:0;font-size:15px;color:{lane['color']};">
                        {lane['label']} <span style="font-size:12px;color:#aaa;float:right;">({lane_count})</span>
                    </h4>
                    <div style="font-size:12px;color:#aaa;margin-top:4px;">
                        Value: <b>{_fmt_inr(lane_rev)}</b>
                        &nbsp;|&nbsp; <span style="color:#f59e0b;">&#9888; At Risk: {_fmt_inr(lane_risk)}</span>
                    </div>
                    <div style="font-size:10px;color:#4b5563;margin-top:6px;border-top:1px solid #1e2235;padding-top:5px;">
                        <span style="color:#6366f1;font-weight:600;">Playbook:</span> {playbook}
                    </div>
                </div>""",
                unsafe_allow_html=True,
            )

            if lane_count == 0:
                st.info("Empty")
                continue

            # ── Cards: only render top 50 per lane for speed ──────────────
            # Pre-compute elite revenue threshold for VIP badge
            _rev_p80_lane = float((lane_df["recent_90_revenue"] * rev_mult).quantile(0.80)) \
                if len(lane_df) > 1 else 0.0
            # Use overall df for elite badge (not lane-local), so top accounts
            # are elite relative to the whole partner base, not just their lane
            _rev_p80_all  = float((df["recent_90_revenue"] * rev_mult).quantile(0.80)) \
                if len(df) > 1 else 0.0

            # Fix D: cap to 30 cards per lane for faster rendering
            shown = lane_df.head(30)
            for _kb_i, (_, row) in enumerate(shown.iterrows()):
                name       = str(row.get("company_name", "Unknown"))
                rev        = _fmt_inr(float(row.get("recent_90_revenue", 0) or 0) * rev_mult)
                rev_raw    = float(row.get("recent_90_revenue", 0) or 0) * rev_mult
                churn_raw  = row.get("churn_probability", 0)
                churn_pct  = f"{float(churn_raw)*100:.0f}%" if pd.notnull(churn_raw) else "—"
                credit     = str(row.get("credit_risk_band", "—"))
                state      = str(row.get("state", "") or "")
                cluster    = str(row.get("cluster_label", "—"))
                at_risk    = row.get("revenue_at_risk", None)
                drop_pct   = row.get("revenue_drop_pct", None)
                recency    = row.get("recency_days", None)
                is_elite   = rev_raw >= _rev_p80_all and _rev_p80_all > 0

                # ── Gone Quiet badge — zero period revenue after prior activity ———
                gone_quiet = bool(row.get("_gone_quiet", False))

                # ── Color-code churn severity ──────────────────────────────
                if pd.notnull(churn_raw) and float(churn_raw) >= 0.7:
                    churn_color = "#ef4444"
                elif pd.notnull(churn_raw) and float(churn_raw) >= 0.5:
                    churn_color = "#f59e0b"
                else:
                    churn_color = "#22c55e"

                # ── Revenue trend display ──────────────────────────────────
                if drop_pct is not None and pd.notnull(drop_pct):
                    dp = float(drop_pct)
                    if dp > 0:
                        rev_trend_html = f"<span style='color:#ef4444;font-weight:600'>↓ {dp:.0f}%</span>"
                    elif dp < 0:
                        # Negative drop = growth
                        rev_trend_html = f"<span style='color:#22c55e;font-weight:600'>↑ {abs(dp):.0f}%</span>"
                    else:
                        rev_trend_html = "<span style='color:#aaa'>→ 0%</span>"
                else:
                    rev_trend_html = "<span style='color:#aaa'>—</span>"

                # ── Last order recency ─────────────────────────────────────
                last_order_str = _days_ago(recency) if recency is not None and pd.notnull(recency) else "—"

                # ── At Risk amount display ─────────────────────────────────
                at_risk_str = _fmt_inr(float(at_risk)) if at_risk is not None and pd.notnull(at_risk) and float(at_risk) > 0 else "—"

                # ── State + cluster + elite badge ─────────────────────────
                cluster_color = "#22c55e" if "VIP" in cluster else "#6366f1" if "Growth" in cluster else "#aaa"
                badge_html = ""
                if gone_quiet:
                    badge_html += (
                        "<span style='background:#1c1000;color:#fbbf24;"
                        "padding:2px 8px;border-radius:4px;font-size:11px;"
                        "font-weight:700;margin-right:4px;border:1px solid #92400e;'"
                        ">&#128274; Gone Quiet</span>"
                    )
                if is_elite:
                    badge_html += (
                        "<span style='background:linear-gradient(90deg,#78350f,#92400e);"
                        "color:#fde68a;padding:2px 8px;border-radius:4px;font-size:11px;"
                        "font-weight:700;margin-right:4px;border:1px solid #b45309;'>"
                        "⭐ Elite Account</span>"
                    )
                if state and state != "—":
                    badge_html += f"<span style='background:#23283a;padding:2px 6px;border-radius:4px;font-size:11px;margin-right:4px;'>📍 {state}</span>"
                if cluster and cluster not in ("—", "nan", "Uncategorized"):
                    badge_html += f"<span style='background:{cluster_color}22;color:{cluster_color};padding:2px 6px;border-radius:4px;font-size:11px;border:1px solid {cluster_color}44;'>🏷 {cluster}</span>"

                # ── Category top spend (from cat_perf_df) ─────────────────
                top_cat_html = ""
                if not cat_perf_df.empty and "category" in cat_perf_df.columns:
                    partner_cats = cat_perf_df[cat_perf_df["company_name"] == name]
                    if not partner_cats.empty:
                        top_cat = partner_cats.loc[partner_cats["total_spend"].idxmax(), "category"]
                        top_cat_spend = partner_cats["total_spend"].max()
                        top_cat_html = (
                            f"<span style='background:#1e1b4b;color:#a5b4fc;padding:2px 6px;"
                            f"border-radius:4px;font-size:11px;margin-left:4px;'>"
                            f"📦 {top_cat} ({_fmt_inr(top_cat_spend)})</span>"
                        )

                elite_prefix = "⭐ " if is_elite else ""
                with st.expander(f"{elite_prefix}{name} — {rev}"):
                    if badge_html or top_cat_html:
                        st.markdown(badge_html + top_cat_html, unsafe_allow_html=True)
                    st.markdown(
                        f"**Rev Trend:** {rev_trend_html} &nbsp;|&nbsp; **Last Order:** {last_order_str}  \n"
                        f"**Churn:** <span style='color:{churn_color};font-weight:600'>{churn_pct}</span>"
                        f" &nbsp;|&nbsp; **Credit:** {credit}  \n"
                        f"**At Risk:** {at_risk_str} &nbsp;|&nbsp; **Cluster:** {cluster}",
                        unsafe_allow_html=True,
                    )
                    # ── Primary risk signal (At Risk / Critical only) ──────────
                    if lane["key"] in ("at_risk", "critical"):
                        sig_text, sig_color = _primary_risk_signal(row)
                        if sig_text:
                            st.markdown(
                                f"<div style='margin-top:8px;padding:6px 10px;"
                                f"background:{sig_color}18;border-radius:6px;"
                                f"border-left:3px solid {sig_color};"
                                f"font-size:12px;color:{sig_color};font-weight:600;'>"
                                f"⚡ {sig_text}"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                    # ── Deep-link button → Partner 360 ─────────────────────────
                    st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)
                    _safe_key = f"kb_jump_{lane['key']}_{_kb_i}_{abs(hash(name)) % 99999}"
                    if st.button(
                        "🔍 Full Report →",
                        key=_safe_key,
                        help=f"Open {name} in Partner 360 View",
                        use_container_width=True,
                    ):
                        st.session_state["preselect_partner"] = name
                        st.session_state["preselect_state"]   = state
                        st.session_state["active_page"]        = "Partner 360 View"
                        st.rerun()
                    _sa_key = f"kb_sa_{lane['key']}_{_kb_i}_{abs(hash(name)) % 99999}"
                    if st.button(
                        "📈 Analyze →",
                        key=_sa_key,
                        help=f"Open {name} in Sales Analyzer",
                        use_container_width=True,
                    ):
                        st.session_state["sa_preselect_partner"] = name
                        st.session_state["active_page"]           = "Sales Analyzer"
                        st.rerun()

            if lane_count > 30:
                st.caption(f"Showing top 30 of {lane_count}. Use filters or search to narrow down.")
