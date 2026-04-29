"""
Sales Analyzer — Extra data helpers + export functions
(YoY, next-order prediction, repeat/new, Excel, PDF, multi-partner)
"""
import io, math
import pandas as pd
from datetime import date, timedelta


# ── YoY comparison ────────────────────────────────────────────────────────────

def fetch_yoy_kpis(engine, party_id: int, start: date, end: date) -> dict:
    span = (end - start).days
    yoy_end   = start - timedelta(days=1)
    yoy_start = yoy_end - timedelta(days=span)
    try:
        df = pd.read_sql(
            """SELECT COALESCE(SUM(tp.net_amt),0) AS revenue,
                      COUNT(DISTINCT t.id)         AS orders
               FROM transactions_dsr t
               JOIN transactions_dsr_products tp ON tp.dsr_id = t.id
               WHERE LOWER(CAST(t.is_approved AS TEXT))='true'
                 AND t.party_id=%(pid)s
                 AND t.date BETWEEN %(s)s AND %(e)s""",
            engine, params={"pid": party_id, "s": yoy_start, "e": yoy_end},
        )
        r = df.iloc[0]
        return {"revenue": float(r["revenue"]), "orders": int(r["orders"]),
                "label": f"{yoy_start.strftime('%d %b %y')} – {yoy_end.strftime('%d %b %y')}"}
    except Exception:
        return {}


# ── Next-order prediction ─────────────────────────────────────────────────────

def predict_next_order(engine, party_id: int) -> dict:
    try:
        df = pd.read_sql(
            """SELECT DISTINCT date FROM transactions_dsr
               WHERE LOWER(CAST(is_approved AS TEXT))='true'
                 AND party_id=%(pid)s ORDER BY date""",
            engine, params={"pid": party_id},
        )
        if len(df) < 3:
            return {}
        dates = pd.to_datetime(df["date"]).sort_values().tolist()
        gaps  = [(dates[i+1]-dates[i]).days for i in range(len(dates)-1)]
        med   = sorted(gaps)[len(gaps)//2]
        last  = dates[-1].date()
        predicted = last + timedelta(days=med)
        days_away = (predicted - date.today()).days
        cv = (pd.Series(gaps).std() / med) if med > 0 else 1
        conf = "High" if cv < 0.4 else "Medium" if cv < 0.8 else "Low"
        return {"predicted_date": predicted, "days_away": days_away,
                "avg_gap": med, "confidence": conf}
    except Exception:
        return {}


# ── Repeat vs New products ────────────────────────────────────────────────────

def fetch_repeat_new(engine, party_id: int, start: date, end: date) -> pd.DataFrame:
    """
    Repeat = product was bought by this partner in ANY transaction OUTSIDE
             the current window (either before OR after, but primarily before).
             This avoids the all-New problem when the window starts at the
             partner's very first order date.
    New    = product appears in this window for the first time ever.
    """
    try:
        return pd.read_sql(
            """SELECT p.product_name,
                      CASE
                        -- Bought BEFORE this window = definite Repeat
                        WHEN EXISTS (
                            SELECT 1 FROM transactions_dsr t2
                            JOIN transactions_dsr_products tp2 ON tp2.dsr_id=t2.id
                            WHERE LOWER(CAST(t2.is_approved AS TEXT))='true'
                              AND t2.party_id=%(pid)s
                              AND tp2.product_id=p.id
                              AND t2.date < %(s)s
                        ) THEN 'Repeat'
                        -- Bought multiple times WITHIN this window = also Repeat
                        WHEN (
                            SELECT COUNT(DISTINCT t3.id)
                            FROM transactions_dsr t3
                            JOIN transactions_dsr_products tp3 ON tp3.dsr_id=t3.id
                            WHERE LOWER(CAST(t3.is_approved AS TEXT))='true'
                              AND t3.party_id=%(pid)s
                              AND tp3.product_id=p.id
                              AND t3.date BETWEEN %(s)s AND %(e)s
                        ) > 1 THEN 'Repeat'
                        ELSE 'New'
                      END AS purchase_type
               FROM master_products p
               WHERE p.id IN (
                   SELECT DISTINCT tp.product_id
                   FROM transactions_dsr t
                   JOIN transactions_dsr_products tp ON tp.dsr_id=t.id
                   WHERE LOWER(CAST(t.is_approved AS TEXT))='true'
                     AND t.party_id=%(pid)s
                     AND t.date BETWEEN %(s)s AND %(e)s
               )""",
            engine, params={"pid": party_id, "s": start, "e": end},
        )
    except Exception:
        return pd.DataFrame()


# ── Multi-partner KPIs ────────────────────────────────────────────────────────

def fetch_multi_partner_kpis(engine, party_ids: list, start: date, end: date) -> pd.DataFrame:
    rows = []
    for pid in party_ids:
        try:
            df = pd.read_sql(
                """SELECT mp.company_name,
                          COALESCE(SUM(tp.net_amt),0)  AS revenue,
                          COUNT(DISTINCT t.id)          AS orders,
                          COALESCE(SUM(tp.qty),0)       AS total_qty,
                          COUNT(DISTINCT tp.product_id) AS unique_products,
                          MAX(t.date)                   AS last_order
                   FROM transactions_dsr t
                   JOIN transactions_dsr_products tp ON tp.dsr_id=t.id
                   JOIN master_party mp ON mp.id=t.party_id
                   WHERE LOWER(CAST(t.is_approved AS TEXT))='true'
                     AND t.party_id=%(pid)s
                     AND t.date BETWEEN %(s)s AND %(e)s
                   GROUP BY mp.company_name""",
                engine, params={"pid": pid, "s": start, "e": end},
            )
            if not df.empty:
                rows.append(df.iloc[0].to_dict())
        except Exception:
            pass
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def fetch_multi_monthly(engine, party_id: int, name: str, start: date, end: date) -> pd.DataFrame:
    try:
        df = pd.read_sql(
            """SELECT DATE_TRUNC('month',t.date)::date AS month,
                      ROUND(SUM(tp.net_amt)::NUMERIC,2)  AS revenue
               FROM transactions_dsr t
               JOIN transactions_dsr_products tp ON tp.dsr_id=t.id
               WHERE LOWER(CAST(t.is_approved AS TEXT))='true'
                 AND t.party_id=%(pid)s
                 AND t.date BETWEEN %(s)s AND %(e)s
               GROUP BY 1 ORDER BY 1""",
            engine, params={"pid": party_id, "s": start, "e": end},
        )
        df["partner"] = name
        return df
    except Exception:
        return pd.DataFrame()


# ── Excel Export ──────────────────────────────────────────────────────────────

def export_excel(partner_name: str, kpis: dict, yoy: dict,
                 prod_df: pd.DataFrame, month_df: pd.DataFrame,
                 sel_start: date, sel_end: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BG   = "0D0F1A"; ACC  = "34D399"; HDR  = "1A1C23"; TXT  = "ECFDF5"; DIM  = "64748B"
    wb   = Workbook()

    def _hfill(c): return PatternFill("solid", fgColor=HDR)
    def _afill(c): return PatternFill("solid", fgColor=ACC)
    def _bgfill():  return PatternFill("solid", fgColor=BG)
    thin = Border(*[Side(style="thin", color="1E293B")]*4)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    c = ws["A1"]; c.value = f"Sales Analyzer — {partner_name}"
    c.fill = PatternFill("solid", fgColor=ACC)
    c.font = Font(bold=True, color=BG, size=14, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"Period: {sel_start} → {sel_end}"
    ws["A2"].font = Font(italic=True, color=DIM, size=10, name="Calibri")
    ws["A2"].fill = _bgfill()

    labels = ["Revenue", "Orders", "Total Qty", "Products", "Last Order",
              "YoY Revenue (prior period)", "YoY Orders (prior period)"]
    values = [
        f"Rs {kpis.get('revenue',0):,.0f}", kpis.get('orders',0),
        f"{kpis.get('total_qty',0):,.0f}", kpis.get('unique_products',0),
        kpis.get('last_order','—'),
        f"Rs {yoy.get('revenue',0):,.0f}" if yoy else "—",
        yoy.get('orders','—') if yoy else "—",
    ]
    for i, (lbl, val) in enumerate(zip(labels, values), start=4):
        ws[f"A{i}"] = lbl; ws[f"B{i}"] = val
        ws[f"A{i}"].fill = _hfill(0); ws[f"B{i}"].fill = _bgfill()
        ws[f"A{i}"].font = Font(color=TXT, size=11, name="Calibri")
        ws[f"B{i}"].font = Font(color=ACC, bold=True, size=11, name="Calibri")
        for col in "AB":
            ws[f"{col}{i}"].border = thin

    # ── Sheet 2: Products ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Products"); ws2.sheet_view.showGridLines = False
    if not prod_df.empty:
        cols = ["product_name","category","product_group","total_qty",
                "avg_rate","total_amount","txn_count","last_purchased"]
        hdrs = ["Product","Category","Group","Qty","Avg Rate","Total Amount","Invoices","Last Purchased"]
        for ci, h in enumerate(hdrs, 1):
            c = ws2.cell(1, ci, h)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.font = Font(bold=True, color=ACC, size=10, name="Calibri")
            c.border = thin; c.alignment = Alignment(horizontal="center")
        for ri, (_, row) in enumerate(prod_df[cols].iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                c = ws2.cell(ri, ci, row[col])
                c.fill = PatternFill("solid", fgColor=BG if ri%2==0 else "12141C")
                c.font = Font(color=TXT, size=10, name="Calibri"); c.border = thin
        for ci in range(1, len(cols)+1):
            ws2.column_dimensions[get_column_letter(ci)].width = 20
        ws2.freeze_panes = "A2"

    # ── Sheet 3: Monthly ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Monthly Revenue"); ws3.sheet_view.showGridLines = False
    if not month_df.empty:
        for ci, h in enumerate(["Month","Revenue","Orders"], 1):
            c = ws3.cell(1, ci, h)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.font = Font(bold=True, color=ACC, size=10, name="Calibri"); c.border = thin
        for ri, (_, row) in enumerate(month_df.iterrows(), 2):
            ws3.cell(ri,1,str(row.get("month",""))).fill = PatternFill("solid",fgColor=BG)
            ws3.cell(ri,2,float(row.get("revenue",0))).fill = PatternFill("solid",fgColor=BG)
            ws3.cell(ri,3,int(row.get("orders",0))).fill = PatternFill("solid",fgColor=BG)
            for ci in range(1,4):
                ws3.cell(ri,ci).font = Font(color=TXT,size=10,name="Calibri"); ws3.cell(ri,ci).border=thin
        for ci in range(1,4):
            ws3.column_dimensions[get_column_letter(ci)].width = 18

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── PDF Export ────────────────────────────────────────────────────────────────



def _pdf_safe(text: str) -> str:
    """Strip/replace characters outside latin-1 range for fpdf Helvetica."""
    repl = {
        '\u2192': '->', '\u2190': '<-', '\u2191': '^', '\u2193': 'v',
        '\u20b9': 'Rs', '\u2022': '-', '\u2013': '-', '\u2014': '-',
        '\u2019': "'", '\u2018': "'", '\u201c': '"', '\u201d': '"',
        '\u2026': '...', '\u00b7': '-', '\u00d7': 'x',
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    return text.encode('latin-1', errors='ignore').decode('latin-1')


def export_pdf(partner_name: str, city: str, state: str,
               kpis: dict, yoy: dict, prod_df: pd.DataFrame,
               sel_start: date, sel_end: date) -> bytes:
    from fpdf import FPDF
    BG=(13,15,26); ACC=(52,211,153); HDR=(26,28,35); TXT=(203,213,225)
    DIM=(100,116,139); WHT=(255,255,255)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.set_margins(12, 12, 12)
    pdf.add_page()

    # Background
    pdf.set_fill_color(*BG); pdf.rect(0,0,297,210,"F")
    # Header bar
    pdf.set_fill_color(*ACC); pdf.rect(0,0,297,24,"F")
    pdf.set_text_color(*BG); pdf.set_font("Helvetica","B",16)
    pdf.set_xy(12,4); pdf.cell(0,10,_pdf_safe(f"Sales Analyzer - {partner_name}"))
    pdf.set_font("Helvetica","",9); pdf.set_xy(12,14)
    pdf.cell(0,6,_pdf_safe(f"{city}, {state} | {sel_start} to {sel_end} | Generated: {date.today()}"))

    # KPI block
    pdf.set_y(30)
    kpi_items = [
        ("Revenue", f"Rs {kpis.get('revenue',0):,.0f}"),
        ("Orders",  str(kpis.get('orders',0))),
        ("Total Qty", f"{kpis.get('total_qty',0):,.0f}"),
        ("Products", str(kpis.get('unique_products',0))),
        ("Last Order", kpis.get('last_order','—')),
        ("YoY Revenue", f"Rs {yoy.get('revenue',0):,.0f}" if yoy else "—"),
    ]
    pdf.set_fill_color(*HDR); pdf.set_text_color(*ACC)
    pdf.set_font("Helvetica","B",10)
    col_w = 44
    for i, (lbl, val) in enumerate(kpi_items):
        x = 12 + i*col_w; pdf.set_xy(x, 30)
        pdf.set_fill_color(*HDR); pdf.cell(col_w-2, 8, _pdf_safe(lbl), fill=True)
        pdf.set_xy(x, 38); pdf.set_text_color(*WHT); pdf.set_font("Helvetica","B",12)
        pdf.cell(col_w-2, 10, _pdf_safe(str(val)), fill=True); pdf.set_font("Helvetica","",10)

    # Product table
    if not prod_df.empty:
        pdf.set_y(55)
        pdf.set_fill_color(*ACC); pdf.set_text_color(*BG); pdf.set_font("Helvetica","B",9)
        col_widths = [80,35,25,30,30,25]
        headers    = ["Product","Category","Qty","Avg Rate","Total Amt","Invoices"]
        for w,h in zip(col_widths,headers):
            pdf.cell(w,7,h,fill=True)
        pdf.ln()
        pdf.set_font("Helvetica","",8)
        for i,(_, row) in enumerate(prod_df.head(25).iterrows()):
            pdf.set_fill_color(*(HDR if i%2==0 else BG))
            pdf.set_text_color(*TXT)
            vals=[_pdf_safe(str(row.get("product_name","") or ""))[:38],
                  _pdf_safe(str(row.get("category","") or ""))[:16],
                  "{:,.0f}".format(float(row.get('total_qty') or 0)),
                  "Rs {:,.0f}".format(float(row.get('avg_rate') or 0)),
                  "Rs {:,.0f}".format(float(row.get('total_amount') or 0)),
                  str(row.get("txn_count") or 0)]
            for w,v in zip(col_widths,vals):
                pdf.cell(w,6,v,fill=True)
            pdf.ln()

    return bytes(pdf.output())
